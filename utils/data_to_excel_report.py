import os
import shutil
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from utils.setting_utils import get_path_from_settings
from PySide6.QtWidgets import QMessageBox, QInputDialog

def _gui_ask_conflict(excel_path, plot_path, title, output_folder):
    """
    Affiche une boîte de dialogue pour résoudre le conflit.
    Renvoie (excel_path, plot_path, final_title) ou lève KeyboardInterrupt si Annuler.
    """
    def paths_for(t):
        return (os.path.join(output_folder, f"{t}.xlsx"),
                os.path.join(output_folder, f"{t}.png"))

    while True:
        box = QMessageBox()
        box.setIcon(QMessageBox.Warning)
        box.setWindowTitle("Conflit de nom de fichier")
        box.setText(
            "Des fichiers existent déjà :\n"
            f"• {excel_path}\n"
            f"• {plot_path}\n\n"
            "Que souhaitez-vous faire ?"
        )
        overwrite = box.addButton("Écraser", QMessageBox.AcceptRole)
        rename    = box.addButton("Renommer…", QMessageBox.ActionRole)
        cancel    = box.addButton("Annuler", QMessageBox.RejectRole)
        box.setDefaultButton(rename)
        box.exec()

        clicked = box.clickedButton()
        if clicked is overwrite:
            return excel_path, plot_path, title

        if clicked is rename:
            while True:
                new_title, ok = QInputDialog.getText(
                    None,
                    "Renommer",
                    "Nouveau nom (sans extension) :"
                )
                if not ok:
                    break
                new_title = new_title.strip()
                if not new_title:
                    continue
                e2, p2 = paths_for(new_title)
                if os.path.exists(e2) or os.path.exists(p2):
                    warn = QMessageBox()
                    warn.setIcon(QMessageBox.Warning)
                    warn.setWindowTitle("Nom déjà pris")
                    warn.setText(
                        f"{e2}\n{p2}\nexiste déjà.\nMerci d’en choisir un autre."
                    )
                    warn.exec()
                    continue
                return e2, p2, new_title

        if clicked is cancel:
            raise KeyboardInterrupt("Opération annulée par l'utilisateur.")
        
def _resolve_output_paths(output_folder, file_title, conflict_strategy="ask"):
    """
    Renvoie (excel_path, plot_path, final_title) en gérant les conflits.
    conflict_strategy: "ask" | "overwrite" | "auto_rename"
    """
    os.makedirs(output_folder, exist_ok=True)

    def paths_for(title):
        return (os.path.join(output_folder, f"{title}.xlsx"),
                os.path.join(output_folder, f"{title}.png"))

    excel_path, plot_path = paths_for(file_title)

    exists = os.path.exists(excel_path) or os.path.exists(plot_path)
    if not exists:
        return excel_path, plot_path, file_title

    if conflict_strategy == "overwrite":
        return excel_path, plot_path, file_title

    if conflict_strategy == "auto_rename":
        i = 1
        while True:
            candidate = f"{file_title}_{i}"
            e, p = paths_for(candidate)
            if not (os.path.exists(e) or os.path.exists(p)):
                return e, p, candidate
            i += 1

    # conflict_strategy == "ask"
    if conflict_strategy == "ask":
        return _gui_ask_conflict(excel_path, plot_path, file_title, output_folder)
        # autre saisie → on redemande

def load_and_filter_data(file_path):
    # Accept None or non-existing path → return empty DataFrames
    if not file_path or not os.path.isfile(file_path):
        empty = pd.DataFrame(columns=['Time', 'Course', 'Force'])
        return empty.copy(), empty.copy()
    data = pd.read_csv(file_path, sep='\t', header=None, names=['Time', 'Course', 'Force'])
    filtered_data = data[data['Force'] >= 1].reset_index(drop=True)
    return data, filtered_data

def compute_additional_columns(data, factor, lever_arm_mm):
    if data is None or data.empty:
        return data  # rien à faire

    # torque = Force [N] * lever_arm [m]
    data = data.copy()
    data['Couple'] = data['Force'] * (lever_arm_mm * 1e-3)
    data['h2'] = factor * data['Course']
    data['h3'] = data['Course'] - data['h2']
    data['rad'] = np.arctan(data['h3'] / lever_arm_mm)
    data['\u00b0'] = np.degrees(data['rad'])
    # normalize to start at zero
    data['Couple'] -= data['Couple'].iloc[0]
    data['\u00b0'] -= data['\u00b0'].iloc[0]
    return data

def find_intersection(data, torque):
    # Safe if columns missing or DF empty
    if (data is None or data.empty
        or 'Couple' not in data.columns or '\u00b0' not in data.columns):
        return np.nan
    hit = data[data['Couple'] >= torque]
    if hit.empty:
        return np.nan
    return hit['\u00b0'].iloc[0]

def export_to_excel_report(
    flexion_path,
    extension_path,
    output_folder,
    file_title,
    test_date,
    brace_type,
    sample_reference,
    speed,
    force_max,
    operator,
    material,
    torque,
    factor,
    lever_arm_mm,
    note=None,  # <— add this
):

    template_path = get_path_from_settings("template_excel")

    new_excel_path, plot_path, final_title = _resolve_output_paths(
        output_folder, file_title, conflict_strategy="ask"
    )

    shutil.copy(template_path, new_excel_path)

    wb = load_workbook(new_excel_path)
    flexion_ws = wb['Flexion']; extension_ws = wb['Extension']; overview_ws = wb['Overview']
    flexion_raw_ws = wb['Raw_data_Flexion']; extension_raw_ws = wb['Raw_data_Extension']

    # --- FLEXION
    flexion_raw_data, flexion_filtered = load_and_filter_data(flexion_path)
    flexion_data = compute_additional_columns(flexion_filtered, factor, lever_arm_mm)
    flexion_intersection = find_intersection(flexion_data, torque)

    if not flexion_raw_data.empty:
        for i, row in flexion_raw_data.iterrows():
            flexion_raw_ws.cell(row=i+3, column=1, value=row['Time'])
            flexion_raw_ws.cell(row=i+3, column=2, value=row['Course'])
            flexion_raw_ws.cell(row=i+3, column=3, value=row['Force'])

    if flexion_data is not None and not flexion_data.empty:
        for i, row in flexion_data.iterrows():
            for col, val in enumerate(row, 1):
                flexion_ws.cell(row=i+3, column=col, value=val)

    # --- EXTENSION
    extension_raw_data, extension_filtered = load_and_filter_data(extension_path)
    extension_data = compute_additional_columns(extension_filtered, factor, lever_arm_mm)
    extension_intersection = find_intersection(extension_data, torque)

    if not extension_raw_data.empty:
        for i, row in extension_raw_data.iterrows():
            extension_raw_ws.cell(row=i+3, column=1, value=row['Time'])
            extension_raw_ws.cell(row=i+3, column=2, value=row['Course'])
            extension_raw_ws.cell(row=i+3, column=3, value=row['Force'])

    if extension_data is not None and not extension_data.empty:
        for i, row in extension_data.iterrows():
            for col, val in enumerate(row, 1):
                extension_ws.cell(row=i+3, column=col, value=val)

    # --- Overview
    overview_ws['I9']  = None if np.isnan(flexion_intersection)   else float(flexion_intersection)
    overview_ws['I10'] = None if np.isnan(extension_intersection) else float(extension_intersection)
    if not np.isnan(flexion_intersection) and flexion_intersection != 0:
        overview_ws['I13'] = float(torque) / float(flexion_intersection)
    else:
        overview_ws['I13'] = None
    if not np.isnan(extension_intersection) and extension_intersection != 0:
        overview_ws['I14'] = float(torque) / float(extension_intersection)
    else:
        overview_ws['I14'] = None
    # (Optional) write note somewhere:
    # overview_ws['C6'] = note
    _set_overview_direct(overview_ws, {
        "C5": material,        # Material
        "D5": brace_type,      # Brace/Brand type
        "F5": sample_reference,# Sample reference
        "I5": test_date,       # Date
        "J5": operator,        # Operator
        "I17": speed,          # Speed
        "I18": force_max,      # Force max
        # Optionally store computed/config values here too if you have fixed cells for them:
        # "I16": torque,       # Torque threshold (if your template reserves it)
        # "I15": lever_arm_mm, # Lever arm (mm) (if reserved)
        # "I19": factor,       # Factor (if reserved)
        # "C6": note,          # Note (if you want a note line)
        # "B3": file_title,    # Title, if you have a title cell
    })
    # --- Plot (always create one so image is present)
    have_flex = flexion_data is not None and not flexion_data.empty
    have_ext  = extension_data is not None and not extension_data.empty

    import matplotlib
    # force non-GUI backend if needed
    try: matplotlib.get_backend()
    except Exception: matplotlib.use("Agg")

    plt.figure(figsize=(8, 5))
    max_x = 0.0
    if have_flex:
        plt.plot(flexion_data['\u00b0'], flexion_data['Couple'], label="Flexion")
        max_x = max(max_x, float(flexion_data['\u00b0'].max()))
    if have_ext:
        plt.plot(extension_data['\u00b0'], extension_data['Couple'], label="Extension")
        max_x = max(max_x, float(extension_data['\u00b0'].max()))
    plt.axhline(y=torque, linestyle="--", linewidth=2, label=f"{torque} Nm")
    if max_x > 0:
        plt.annotate(f"{torque} Nm", xy=(max_x, torque),
                     xytext=(max_x + 0.15, torque), color="red", fontsize=10, ha="left", va="center")
    if not np.isnan(flexion_intersection):
        plt.axvline(x=float(flexion_intersection), linestyle="--", linewidth=1)
    if not np.isnan(extension_intersection):
        plt.axvline(x=float(extension_intersection), linestyle="--", linewidth=1)
    if not (have_flex or have_ext):
        ax = plt.gca()
        ax.text(0.5, 0.5, "No valid data after filtering", ha="center", va="center",
                transform=ax.transAxes)
        ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    plt.xlim(left=0); plt.ylim(bottom=0)
    plt.title(final_title); plt.xlabel("Angular deflection [\u00b0]"); plt.ylabel("Torque [Nm]")
    plt.legend(); plt.grid(True)
    plt.savefig(plot_path, dpi=150, bbox_inches="tight")
    plt.close()

    if os.path.isfile(plot_path):
        try:
            img = Image(plot_path)
            img.width, img.height = 500, 300
            overview_ws.add_image(img, "B7")
        except Exception as ex:
            print(f"[CORE] Could not insert image: {ex}")

    wb.save(new_excel_path)
    wb.close()

    mechanical_results = {
        "torque_threshold_Nm": torque,
        "angular_deflection_deg": {},
        "rigidity_Nm_per_deg": {}
    }
    if not np.isnan(flexion_intersection) and flexion_intersection != 0:
        mechanical_results["angular_deflection_deg"]["flexion"] = round(float(flexion_intersection), 3)
        mechanical_results["rigidity_Nm_per_deg"]["flexion"]    = round(float(torque) / float(flexion_intersection), 3)
    if not np.isnan(extension_intersection) and extension_intersection != 0:
        mechanical_results["angular_deflection_deg"]["extension"] = round(float(extension_intersection), 3)
        mechanical_results["rigidity_Nm_per_deg"]["extension"]    = round(float(torque) / float(extension_intersection), 3)

    return mechanical_results

def _set_overview_direct(ws, mapping: dict):
    """
    Direct cell writes (if your template has fixed cells).
    mapping: {"A1": value, "B2": value, ...}
    """
    for addr, val in mapping.items():
        try:
            ws[addr] = val
        except Exception:
            pass